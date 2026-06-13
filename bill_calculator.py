from openpyxl import load_workbook, Workbook
import re
from datetime import datetime
file_location = input("Specify the file absolute path:\n")
file = load_workbook(filename = file_location)
ws = file.active

#Prices:
#Assumptions: 
#All backup is the same price
#Snapshots cost as much as the virtual disks
#'CCE PM Cluster(High Availability)|1-10node' is the only expected find for the CCE service
daily_price = {
	'vCPU'	: ,
	'vRAM'	: 0.0 /30,
	'EVS'	: 0.0 / 30,
	'backup': 0.0 / 30,
	'ELB'	: 0.0 / 30,
	'EIP'	: 0.0 / 30,
	'CCE'	: 0.0/30,
	'HSS_enterprise' : 0.0 / 30,
	'HSS_premium'	 : 0.0 / 30,
	'HSS_container'	 : 0.0 / 30,
	'HSS_antitamper' : 0.0 / 30,
	'WAF_100' : 210 / 30.0,
	'WAF_500' : 0.0 / 30,
	'EdgeFW' : 0.0 / 30,
	'OBS' : 0.0 / 30,
	'SFS' : 0.0 / 30
}

def get_col(ws, pattern):
	for col in range(1, ws.max_column + 1):
		cell_value = ws.cell(row=1, column=col).value
		if cell_value and re.search(pattern, cell_value, re.IGNORECASE):
			return col
	return None

def get_time(begin_time, end_time):
	begin_time = datetime.strptime(str(begin_time), '%Y-%m-%d %H:%M:%S')
	end_time =  datetime.strptime(str(end_time), '%Y-%m-%d %H:%M:%S')
	return (end_time - begin_time).total_seconds() /(24 * 3600) #Return time in days

def get_resource_price(product):
	#product_col = get_col(ws, r'Metering\s+Metric')
	#product = ws.cell(row, product_col).value
	#product = product.strip()
	match product:
		case prod if re.match(r'^ECS', prod, re.IGNORECASE):
			#import pdb; pdb.set_trace();
			m = re.search(r'(\d+)CPU_(\d+)(?:GRAM|RAM)', prod, re.IGNORECASE)
			if m:
				cpu_num = int(m.group(1))
				ram_num = int(m.group(2))
				cpu_price = cpu_num * daily_price.get('vCPU', 0.0)
				ram_price = ram_num * daily_price.get('vRAM', 0.0)
				return cpu_price+ram_price

		case prod if re.match(r'^Backup', prod, re.IGNORECASE):
			return daily_price.get('backup', 0.0)

		case prod if re.match(r'^EIP$', prod, re.IGNORECASE): #This is EdgeFW
			return daily_price.get('EdgeFW', 0.0)
		case prod if re.match(r'EIP-Instance', prod, re.IGNORECASE):
			return daily_price.get('EIP',0.0)
		case prod if re.match(r'ELB-Instance', prod, re.IGNORECASE):
			return daily_price.get('ELB', 0.0)
		case prod if re.match(r'EVS.*', prod, re.IGNORECASE):
			return daily_price.get('EVS', 0.0)
		case prod if re.match(r'^HSS', prod, re.IGNORECASE):
			#import pdb; pdb.set_trace();
			m = re.search(r'HSS-HSS(\w+)', prod, re.IGNORECASE)
			if m:
				suffix = m.group(1).lower()
				return daily_price.get(f'HSS_{suffix}', 0.0)
		case prod if re.match(r'.*obs.*', prod, re.IGNORECASE):
			return daily_price.get('OBS', 0.0)
		case prod if re.match(r'^Scalable\s+File\s+Service', prod, re.IGNORECASE):
			return daily_price.get('SFS', 0.0)
		case prod if re.match(r'^Premium\s+WAF.*', prod, re.IGNORECASE):
			m = re.search(r'Premium\s+WAF\((\d+)\)', prod, re.IGNORECASE)
			if m:
				suffix = m.group(1)
				return daily_price.get(f'WAF_{suffix}', 0.0)
		case prod if re.match(r'^CCE.*', prod, re.IGNORECASE):
			return daily_price.get('CCE', 0.0)
		case _:
			print('Error -- row number: ',row)
			pass

#Get the needed columns:
begin_time_col 	   = get_col(ws, r'^Meter\s*Begin\s*Time.*')
print(begin_time_col)
end_time_col   	   = get_col(ws, r'^Meter\s*End\s*Time.*')
product_col	   	   = get_col(ws, r'^Metering\s+Metric')
metering_value_col = get_col(ws, r'^Metering\s*Value')
unit_price_col = get_col(ws,r'^Unit\s*Price.*')
fee_col = get_col(ws,r'^Fee.*')

#Pour them into arrays:
begin_time_arr = [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=begin_time_col, max_col=begin_time_col, values_only=True)]
end_time_arr = [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=end_time_col, max_col=end_time_col, values_only=True)]
product_arr	= [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=product_col, max_col=product_col, values_only=True)]
metering_value_arr	= [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=metering_value_col, max_col=metering_value_col, values_only=True)]
unit_price_arr = []
fee_arr = []
for begin, end, product, meter in zip(begin_time_arr, end_time_arr, product_arr, metering_value_arr):
	total_time = get_time(begin, end)
	unit_price = get_resource_price(product)
	unit_price_arr.append(unit_price)
	fee_arr.append(total_time * unit_price * meter)

for i in range(len(unit_price_arr)):
	excel_row = i + 2
	ws.cell(row=excel_row, column=unit_price_col).value = unit_price_arr[i]
	ws.cell(row=excel_row, column=fee_col).value = fee_arr[i]
file.save('result.xlsx')